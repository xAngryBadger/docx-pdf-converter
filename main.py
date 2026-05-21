from fastapi import FastAPI, UploadFile, File, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse
import os
import uuid
from pathlib import Path
from docx import Document
from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from openpyxl import load_workbook
from pypdf import PdfReader, PdfWriter

FONT_DIR = "/usr/share/fonts/TTF"
for name, file in [
    ("DejaVu", "DejaVuSans.ttf"),
    ("DejaVu-Bold", "DejaVuSans-Bold.ttf"),
    ("DejaVuMono", "DejaVuSansMono.ttf"),
]:
    path = os.path.join(FONT_DIR, file)
    if os.path.exists(path):
        pdfmetrics.registerFont(TTFont(name, path))

app = FastAPI()

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

TEMP_DIR = Path("/tmp/docx-converter")
TEMP_DIR.mkdir(exist_ok=True)

def _get_styles():
    base = getSampleStyleSheet()
    font = "DejaVu" if "DejaVu" in pdfmetrics._fonts else "Helvetica"
    bold = "DejaVu-Bold" if "DejaVu-Bold" in pdfmetrics._fonts else "Helvetica-Bold"
    mono = "DejaVuMono" if "DejaVuMono" in pdfmetrics._fonts else "Courier"
    styles = {
        "Normal": ParagraphStyle("DejaVuNormal", parent=base["Normal"], fontName=font, fontSize=10, leading=14),
        "Heading1": ParagraphStyle("DejaVuH1", parent=base["Heading1"], fontName=bold, fontSize=14, leading=18, spaceAfter=6),
        "Code": ParagraphStyle("DejaVuCode", parent=base["Code"], fontName=mono, fontSize=8, leading=11),
    }
    return styles

def convert_docx_to_pdf(docx_path: str, output_path: str):
    doc = Document(docx_path)
    doc_pdf = SimpleDocTemplate(output_path, pagesize=A4)
    styles = _get_styles()
    story = []

    for para in doc.paragraphs:
        story.append(Paragraph(para.text, styles['Normal']))
        story.append(Spacer(1, 6))

    doc_pdf.build(story)

def convert_xlsx_to_pdf(xlsx_path: str, output_path: str):
    wb = load_workbook(xlsx_path, data_only=True)
    doc_pdf = SimpleDocTemplate(output_path, pagesize=A4)
    styles = _get_styles()
    story = []

    for sheet in wb.worksheets:
        story.append(Paragraph(f"Sheet: {sheet.title}", styles['Heading1']))
        story.append(Spacer(1, 6))

        max_row = sheet.max_row or 0
        max_col = sheet.max_column or 0

        for row in sheet.iter_rows(min_row=1, max_row=min(50, max_row), min_col=1, max_col=min(20, max_col)):
            row_text = " | ".join(str(cell.value) if cell.value else "" for cell in row)
            story.append(Paragraph(row_text, styles['Code']))
            story.append(Spacer(1, 3))

        story.append(Spacer(1, 12))

    doc_pdf.build(story)

def compress_pdf_file(input_path: str, output_path: str):
    reader = PdfReader(input_path)
    writer = PdfWriter()

    for page in reader.pages:
        writer.add_page(page)

    with open(output_path, "wb") as f:
        writer.write(f)

@app.post("/api/convert/docx-pdf")
async def convert_docx_pdf(file: UploadFile = File(...)):
    try:
        uid = uuid.uuid4().hex[:8]
        temp_input = TEMP_DIR / f"input_{uid}_{file.filename}"
        temp_output = TEMP_DIR / f"output_{uid}_{file.filename.replace('.docx', '.pdf')}"

        with open(temp_input, "wb") as f:
            content = await file.read()
            f.write(content)

        convert_docx_to_pdf(str(temp_input), str(temp_output))

        return FileResponse(
            str(temp_output),
            media_type="application/pdf",
            filename=temp_output.name
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/convert/xlsx-pdf")
async def convert_xlsx_pdf(file: UploadFile = File(...)):
    try:
        uid = uuid.uuid4().hex[:8]
        temp_input = TEMP_DIR / f"input_{uid}_{file.filename}"
        temp_output = TEMP_DIR / f"output_{uid}_{file.filename.replace('.xlsx', '.pdf')}"

        with open(temp_input, "wb") as f:
            content = await file.read()
            f.write(content)

        convert_xlsx_to_pdf(str(temp_input), str(temp_output))

        return FileResponse(
            str(temp_output),
            media_type="application/pdf",
            filename=temp_output.name
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/convert/pdf-compress")
async def compress_pdf(file: UploadFile = File(...)):
    try:
        uid = uuid.uuid4().hex[:8]
        temp_input = TEMP_DIR / f"input_{uid}_{file.filename}"
        temp_output = TEMP_DIR / f"compressed_{uid}_{file.filename}"

        with open(temp_input, "wb") as f:
            content = await file.read()
            f.write(content)

        compress_pdf_file(str(temp_input), str(temp_output))

        return FileResponse(
            str(temp_output),
            media_type="application/pdf",
            filename=temp_output.name
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/convert/batch")
async def convert_batch(files: list[UploadFile] = File(...)):
    try:
        results = []
        for file in files:
            try:
                uid = uuid.uuid4().hex[:8]
                temp_input = TEMP_DIR / f"input_{uid}_{file.filename}"

                with open(temp_input, "wb") as f:
                    content = await file.read()
                    f.write(content)

                if file.filename.endswith('.docx'):
                    temp_output = TEMP_DIR / f"output_{uid}_{file.filename.replace('.docx', '.pdf')}"
                    convert_docx_to_pdf(str(temp_input), str(temp_output))
                elif file.filename.endswith('.xlsx'):
                    temp_output = TEMP_DIR / f"output_{uid}_{file.filename.replace('.xlsx', '.pdf')}"
                    convert_xlsx_to_pdf(str(temp_input), str(temp_output))
                else:
                    results.append({
                        "original": file.filename,
                        "status": "error",
                        "error": "Formato não suportado no modo lote. Use DOCX ou XLSX."
                    })
                    continue

                results.append({
                    "original": file.filename,
                    "converted": temp_output.name,
                    "status": "success"
                })
            except Exception as e:
                results.append({
                    "original": file.filename,
                    "status": "error",
                    "error": str(e)
                })

        return {"results": results}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/")
async def root():
    return {"message": "DocX-PDF Converter API"}

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=int(os.environ.get("PORT", 8001)))
